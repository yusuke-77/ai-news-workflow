"""
Step 2: 収集した AI ニュースを Excel に保存する
既存ファイルがあればデータ行だけ上書きし、書式は維持する。
"""
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.dirname(__file__)
EXCEL_PATH = os.path.join(BASE_DIR, "output", "ai_news.xlsx")

HEADERS    = ["No", "タイトル", "概要", "URL", "公開日時", "ソース", "収集日時"]
COL_WIDTHS = [5,    55,         65,    45,    18,         20,       20]

COLOR_HEADER_BG  = "1F3864"
COLOR_HEADER_FG  = "FFFFFF"
COLOR_ROW_EVEN   = "EEF3FB"


def _apply_header(ws) -> None:
    fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    font = Font(bold=True, color=COLOR_HEADER_FG, size=11)
    align = Alignment(horizontal="center", vertical="center")

    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 26


def save_to_excel(news_items: list[dict]) -> str:
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["AIニュース"] if "AIニュース" in wb.sheetnames else wb.create_sheet("AIニュース")
        # ヘッダー行だけ残してデータ行を空にする
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
                cell.hyperlink = None
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AIニュース"
        _apply_header(ws)

    collected_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    even_fill = PatternFill(start_color=COLOR_ROW_EVEN, end_color=COLOR_ROW_EVEN, fill_type="solid")

    for i, item in enumerate(news_items, start=2):
        values = [
            i - 1,
            item["title"],
            item["summary"],
            item["url"],
            item["published"].strftime("%Y-%m-%d %H:%M"),
            item["source"],
            collected_at,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if i % 2 == 0:
                cell.fill = even_fill

    wb.save(EXCEL_PATH)
    print(f"  Excel 保存完了: {EXCEL_PATH}")
    return EXCEL_PATH


if __name__ == "__main__":
    from collect_news import collect_news
    save_to_excel(collect_news())
